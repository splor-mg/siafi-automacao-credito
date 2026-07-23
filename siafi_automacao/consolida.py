import os
import sys
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv

load_dotenv()
BASE = os.getenv('ONEDRIVE_BASE')

# ---------------------------------------------------------------------------
# Caminhos
# ---------------------------------------------------------------------------
DIR_SCRIPTS = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.realpath(os.path.join(DIR_SCRIPTS, '..', 'data'))

# modelo.xlsm  -> versionado no Git, nunca é escrito (macros + aba ROBO)
# copia.xlsm   -> gerado a cada execução a partir do modelo, fora do Git
MODELO_PATH = os.path.join(DATA_DIR, 'modelo.xlsm')
XLSM_PATH = os.path.join(DATA_DIR, 'copia.xlsm')

SHEET_ENTRADA = 'PREENCHER AQUI'

folder_paths = [
    os.path.join(BASE, 'Remanejamentos'),
    os.path.join(BASE, 'Remanejamentos (SEGOV_DCAF_FONTE 95_ FONTE 80)'),
]
output_file = os.path.join(BASE, 'Robo', 'copia.xlsx')
processed_folder = os.path.join(BASE, 'Realizados')


# ---------------------------------------------------------------------------
# Auxiliares
# ---------------------------------------------------------------------------

def abortar(titulo, mensagens):
    """Imprime o bloco de erro e encerra com código 1 (tratado pelo login.py)."""
    print()
    print("=" * 70)
    print(f"  {titulo}")
    print("=" * 70)
    for m in mensagens:
        print(f"  - {m}")
    print("=" * 70)
    print()
    sys.exit(1)


def destino_sem_colisao(pasta, filename):
    """Devolve um caminho de destino livre, acrescentando (1), (2)... se preciso."""
    destino = os.path.join(pasta, filename)
    if not os.path.exists(destino):
        return destino

    nome, ext = os.path.splitext(filename)
    contador = 1
    while os.path.exists(destino):
        destino = os.path.join(pasta, f"{nome} ({contador}){ext}")
        contador += 1
    return destino


# ---------------------------------------------------------------------------
# Fluxo principal
# ---------------------------------------------------------------------------

def combine_excel_files(folder_paths, output_file, processed_folder):
    if not os.path.exists(MODELO_PATH):
        abortar(
            "MODELO NÃO ENCONTRADO",
            [
                f"Esperado em: {MODELO_PATH}",
                "O modelo.xlsm é versionado no repositório e contém as macros e a aba ROBO.",
                "Rode 'git checkout -- data/modelo.xlsm' para restaurá-lo.",
            ],
        )

    os.makedirs(processed_folder, exist_ok=True)
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # -----------------------------------------------------------------------
    # 1. Leitura de TODOS os arquivos antes de mover qualquer um
    # -----------------------------------------------------------------------
    erros = []
    lidos = []  # (file_path, filename, df)

    for folder_path in folder_paths:
        if not os.path.isdir(folder_path):
            erros.append(f"Pasta não encontrada: {folder_path}")
            continue

        xlsm_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsm')]

        for filename in sorted(xlsm_files):
            file_path = os.path.join(folder_path, filename)
            try:
                df = pd.read_excel(file_path, sheet_name=SHEET_ENTRADA)
            except Exception as e:
                erros.append(f"{filename}: {e}")
                continue

            lidos.append((file_path, filename, df))
            print(f"Arquivo {filename} lido com sucesso.")

    if erros:
        abortar(
            "ERROS NA LEITURA DAS PLANILHAS — NENHUM ARQUIVO FOI MOVIDO",
            erros + [f"Verifique se todos os arquivos possuem a aba '{SHEET_ENTRADA}'."],
        )

    if not lidos:
        abortar(
            "NENHUMA PLANILHA ENCONTRADA PARA PROCESSAR",
            [f"Pastas verificadas: {', '.join(folder_paths)}"],
        )

    # -----------------------------------------------------------------------
    # 2. Consolidação
    # -----------------------------------------------------------------------
    combined_df = pd.concat([df for _, _, df in lidos], ignore_index=True)
    combined_df = combined_df.dropna(how='all')
    combined_df = combined_df.sort_values(
        by=['TIPO', 'UO_COD', 'ORIENTACAO'],
        ascending=[True, True, False],
    )
    combined_df = combined_df.reset_index(drop=True)

    if combined_df.empty:
        abortar(
            "PLANILHAS LIDAS, MAS SEM LINHAS DE DADOS",
            [f"{len(lidos)} arquivo(s) lido(s), nenhuma linha preenchida encontrada."],
        )

    # -----------------------------------------------------------------------
    # 3. Cópia de conferência em .xlsx no OneDrive
    # -----------------------------------------------------------------------
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name=SHEET_ENTRADA, index=False)

    print(f"Cópia .xlsx salva em: {output_file}")

    # -----------------------------------------------------------------------
    # 4. Geração do copia.xlsm a partir do modelo (preserva macros e aba ROBO)
    # -----------------------------------------------------------------------
    if os.path.exists(XLSM_PATH):
        os.remove(XLSM_PATH)
    shutil.copyfile(MODELO_PATH, XLSM_PATH)

    wb = load_workbook(XLSM_PATH, keep_vba=True)
    ws = wb[SHEET_ENTRADA]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for r_idx, row_data in enumerate(
        dataframe_to_rows(combined_df, index=False, header=False), start=2
    ):
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(XLSM_PATH)
    print(f"Dados escritos em copia.xlsm ({len(combined_df)} linhas).")

    # -----------------------------------------------------------------------
    # 5. Só agora move os arquivos de origem para 'Realizados'
    # -----------------------------------------------------------------------
    for file_path, filename, _ in lidos:
        destino = destino_sem_colisao(processed_folder, filename)
        shutil.move(file_path, destino)
        print(f"Arquivo movido para: {destino}")


if __name__ == '__main__':
    combine_excel_files(
        folder_paths=folder_paths,
        output_file=output_file,
        processed_folder=processed_folder,
    )