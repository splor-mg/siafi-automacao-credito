"""
Registro do desfecho de cada documento enviado ao SIAFI, gravado na aba ROBO.

Quem sabe o desfecho (registro efetuado / inconsistência) e o número do
documento é o finalizar_documento; quem sabe a que linha da planilha aquilo
corresponde é o login.py. Este módulo liga as duas pontas sem alterar nenhuma
assinatura: o login.py chama marcar_linha() ao fim de cada iteração e o
finalizar_documento chama registrar_ok()/registrar_erro().

Um documento agrupa várias linhas da planilha e só é encerrado quando a linha
seguinte muda de UO ou de TIPO. Logo, a "última linha processada" no momento em
que o finalizar_documento roda é exatamente a última linha que entrou no
documento sendo encerrado — que é onde a conferência espera o 'ok'.
"""

import os

from openpyxl import load_workbook

SHEET_NAME = 'ROBO'
COL_PROGRESSO = 21  # coluna U
COL_SIAFI = 22      # coluna V

_ultima_linha = None
_resultados = []  # (linha_excel, uo, status, nr_doc)


def marcar_linha(linha_excel):
    """Registra qual foi a última linha da planilha processada."""
    global _ultima_linha
    _ultima_linha = linha_excel


def registrar_ok(uo, nr_doc):
    _registrar(uo, 'ok', nr_doc)


def registrar_erro(uo):
    _registrar(uo, 'erro', '')


def _registrar(uo, status, nr_doc):
    if _ultima_linha is None:
        # Documento encerrado antes de qualquer linha ter sido processada:
        # não há onde gravar, mas o terminal já informou o desfecho.
        return
    _resultados.append((_ultima_linha, uo, status, nr_doc))


def houve_erro():
    return any(status == 'erro' for _, _, status, _ in _resultados)


def gravar(caminho_xlsm):
    """Escreve 'ok'/'erro' na coluna U e o nº do documento na coluna V."""
    if not _resultados:
        print("Nenhum documento concluído — nada a gravar na planilha.")
        return

    # Roda dentro do finally do login.py: uma falha aqui (planilha aberta no
    # Excel, por exemplo) não pode derrubar a cópia de conferência nem esconder
    # o erro que interrompeu o fluxo. O resumo no terminal é o plano B.
    try:
        wb = load_workbook(caminho_xlsm, keep_vba=True)
        ws = wb[SHEET_NAME]

        for linha_excel, _uo, status, nr_doc in _resultados:
            ws.cell(linha_excel, COL_PROGRESSO).value = status
            if status == 'ok':
                # texto, para preservar os zeros à esquerda do nº do documento
                ws.cell(linha_excel, COL_SIAFI).value = str(nr_doc)

        if wb.calculation is not None:
            # A aba ROBO é toda fórmula: sem o cache do Excel, força o recálculo
            # na abertura para que A–T voltem a exibir os valores.
            wb.calculation.fullCalcOnLoad = True

        wb.save(caminho_xlsm)
    except Exception as e:
        print()
        print("=" * 70)
        print("  NÃO FOI POSSÍVEL GRAVAR O RESULTADO NA PLANILHA")
        print("=" * 70)
        print(f"  Arquivo: {caminho_xlsm}")
        print(f"  Motivo: {e}")
        print("  Feche o arquivo no Excel e anote os dados do resumo abaixo.")
        print("=" * 70)
        return

    print(f"Resultados gravados em {os.path.basename(caminho_xlsm)} "
          f"({len(_resultados)} documento(s)).")


def imprimir_resumo():
    if not _resultados:
        return

    erros = [r for r in _resultados if r[2] == 'erro']

    print()
    print("-" * 70)
    print(f"Documentos concluídos: {len(_resultados)} | "
          f"OK: {len(_resultados) - len(erros)} | Com erro: {len(erros)}")
    for linha_excel, uo, status, nr_doc in _resultados:
        detalhe = f"nº {nr_doc}" if status == 'ok' else "não registrado no SIAFI"
        print(f"  linha {linha_excel} | UO {uo} | {status.upper()} | {detalhe}")
    print("-" * 70)

    if erros:
        print()
        print("=" * 70)
        print("  ATENÇÃO — DOCUMENTOS RECUSADOS PELO SIAFI")
        print("=" * 70)
        for linha_excel, uo, _status, _nr in erros:
            print(f"  UO {uo} (linha {linha_excel} da planilha)")
        print("  Confira essas UOs no SIAFI antes de reprocessar.")
        print("=" * 70)
